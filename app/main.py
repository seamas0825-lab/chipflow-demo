"""
ChipFlow FastAPI Main Application.
Provides REST and Server-Sent Events (SSE) APIs for BOM parsing, Ego-style price checking,
Vector alternative recommendation, European cross-border sourcing, and AI Chat.
Compatible with local execution and Vercel Serverless runtime.
"""

import io
import os
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import openpyxl

from app.config import settings
from app.data.sample_boms import SAMPLE_BOMS
from app.data.european_inventory import EUROPEAN_DEADSTOCK_INVENTORY
from app.services.bom_parser import parse_excel_bom, parse_text_bom
from app.services.pricing_agent import stream_price_lookup, get_cached_price, LIVE_PRICE_CACHE
from app.services.vector_matcher import find_vector_substitutes
from app.services.crossborder_sourcing import generate_trilingual_inquiry_email, check_european_deadstock
from app.services.compliance_checker import check_export_compliance
from app.services.bom_ai_chat import answer_bom_query

app = FastAPI(title=settings.APP_NAME, version=settings.VERSION)

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")
INDEX_HTML_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates", "index.html")

if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/", response_class=HTMLResponse)
async def serve_index(request: Request):
    try:
        with open(INDEX_HTML_PATH, "r", encoding="utf-8") as f:
            html = f.read()
        html = html.replace("{{ app_name }}", settings.APP_NAME).replace("{{ version }}", settings.VERSION)
        return HTMLResponse(content=html)
    except Exception as e:
        return HTMLResponse(content=f"<h1>Error loading index.html: {e}</h1>", status_code=500)

@app.get("/api/sample-boms")
async def get_sample_boms():
    return {
        "success": True,
        "data": [
            {"id": k, "name": v["name"], "description": v["description"], "count": len(v["items"])}
            for k, v in SAMPLE_BOMS.items()
        ]
    }

@app.get("/api/sample-boms/{bom_id}")
async def get_single_sample_bom(bom_id: str):
    if bom_id not in SAMPLE_BOMS:
        raise HTTPException(status_code=404, detail="BOM not found")
    bom = SAMPLE_BOMS[bom_id]
    
    enriched_items = []
    for it in bom["items"]:
        mpn = it["mpn"]
        price_info = get_cached_price(mpn)
        eu_info = check_european_deadstock(mpn)
        compliance = check_export_compliance(mpn, it["manufacturer"], it["category"])
        
        enriched_items.append({
            **it,
            "quote": price_info,
            "has_european_stock": eu_info.get("matched", False),
            "european_stock_info": eu_info.get("inventory") if eu_info.get("matched") else None,
            "compliance": compliance
        })
        
    return {
        "success": True,
        "bom_id": bom["id"],
        "name": bom["name"],
        "description": bom["description"],
        "items": enriched_items
    }

@app.post("/api/upload-bom")
async def upload_bom(
    file: UploadFile = File(None),
    raw_text: str = Form(None)
):
    if file and file.filename:
        content = await file.read()
        items = parse_excel_bom(content)
    elif raw_text:
        items = parse_text_bom(raw_text)
    else:
        raise HTTPException(status_code=400, detail="请上传 Excel 文件或输入 BOM 文本")
        
    enriched_items = []
    for it in items:
        mpn = it["mpn"]
        price_info = get_cached_price(mpn)
        eu_info = check_european_deadstock(mpn)
        compliance = check_export_compliance(mpn, it["manufacturer"], it["category"])
        
        enriched_items.append({
            **it,
            "quote": price_info,
            "has_european_stock": eu_info.get("matched", False),
            "european_stock_info": eu_info.get("inventory") if eu_info.get("matched") else None,
            "compliance": compliance
        })
        
    return {
        "success": True,
        "filename": file.filename if file else "Custom Text BOM",
        "items": enriched_items
    }

@app.get("/api/stream-quote/{mpn}")
async def stream_quote(mpn: str, qty: int = 1000):
    return StreamingResponse(
        stream_price_lookup(mpn, qty),
        media_type="text/event-stream"
    )

@app.get("/api/vector-substitutes/{mpn}")
async def get_vector_substitutes(mpn: str):
    data = find_vector_substitutes(mpn)
    return {"success": True, "data": data}

@app.get("/api/european-stock/{mpn}")
async def get_european_stock(mpn: str):
    data = check_european_deadstock(mpn)
    return {"success": True, "data": data}

@app.post("/api/generate-email")
async def create_email(payload: dict):
    mpn = payload.get("mpn", "")
    target_qty = payload.get("quantity", 1000)
    company_name = payload.get("company_name", "ChipFlow Cross-Border Sourcing Hub")
    
    result = generate_trilingual_inquiry_email(mpn, target_qty, company_name)
    return {"success": True, "data": result}

@app.post("/api/chat-bom")
async def chat_bom(payload: dict):
    query = payload.get("query", "")
    items = payload.get("items", [])
    reply = answer_bom_query(query, items)
    return {"success": True, "reply": reply}

@app.post("/api/export-quote-excel")
async def export_quote_excel(payload: dict):
    items = payload.get("items", [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ChipFlow 智能报价汇总"
    
    headers = [
        "行号", "物料型号 (MPN)", "品牌", "封装", "需求数量", 
        "目标单价 (¥)", "推荐报价 (¥)", "推荐货源渠道", "库存状态", "交期", "预计总价 (¥)", "合规等级"
    ]
    ws.append(headers)
    
    # Header styling
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        
    for it in items:
        quote = it.get("quote", {}).get("best_spot", {})
        unit_price = quote.get("price_cny", it.get("target_price_cny", 0))
        qty = it.get("quantity", 1000)
        total_price = round(unit_price * qty, 2)
        
        ws.append([
            it.get("line_no"),
            it.get("mpn"),
            it.get("manufacturer"),
            it.get("package"),
            qty,
            it.get("target_price_cny", 0),
            unit_price,
            quote.get("source", "多渠道汇总"),
            "充足" if quote.get("stock", 0) > 0 else "紧缺/替代",
            quote.get("lead_time", "3-5天"),
            total_price,
            it.get("compliance", {}).get("eccn", "EAR99")
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ChipFlow_Quotation_Matrix.xlsx"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=8090, reload=True)
